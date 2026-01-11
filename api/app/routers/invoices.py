# api/app/routers/invoices.py
from datetime import datetime, date
from typing import Optional, List, Dict, Any, Set
from decimal import Decimal, InvalidOperation
import csv, io, json, re, statistics
from math import ceil

from sqlalchemy import func, and_, or_, case
from ..shared import APIRouter, Depends, HTTPException, BaseModel, Field, Session
from ..database import get_db
from ..models import Invoice, Customer, InvoiceUploadPreset, Payment, PaymentAllocation, User
from ..routers.auth import require_user
from ..calculate_due_date import compute_due_date
from io import BytesIO
from openpyxl import load_workbook
from fastapi import UploadFile, File, Form
from sqlalchemy.exc import IntegrityError

router = APIRouter(prefix="/api/invoices", tags=["invoices"])

# ----------------------------
# Pydantic models
# ----------------------------

class InvoiceIn(BaseModel):
    customer_id: int
    invoice_number: str
    amount_due: Decimal
    currency: str = "GBP"
    issue_date: str | None = None
    due_date: str | None = None
    terms_type: str | None = None
    terms_days: int | None = None

class InvoiceOut(BaseModel):
    id: int
    customer_id: int
    invoice_number: str
    amount_due: Decimal
    status: str
    due_date: Optional[str]
    reconciliation_ref: Optional[str]

class UploadPresetIn(BaseModel):
    name: str
    mapping: Dict[str, Any]
    header: bool = True
    delimiter: str = ","
    date_format: Optional[str] = None
    default_customer_id: Optional[int] = None
    assign_mode: str = "per_row"            # "per_row" | "single"
    create_missing_customers: bool = False

class UploadPresetOut(BaseModel):
    id: int
    name: str
    mapping: Dict[str, Any]
    header: bool
    delimiter: str
    date_format: Optional[str]
    default_customer_id: Optional[int]
    assign_mode: str
    create_missing_customers: bool

class InvoiceRow(BaseModel):
    id: int
    customer_id: int
    customer_name: str
    invoice_number: Optional[str]
    issue_date: Optional[str]
    due_date: Optional[str]
    amount: float          # original amount_due
    remaining: float       # amount - allocations
    status: str            # open|overdue|paid
    days_overdue: int

class PageOut(BaseModel):
    items: List[InvoiceRow]
    page: int
    per_page: int
    total: int
    pages: int

# ----------------------------
# helper 
# ----------------------------

def _recalc_invoice_paid_fields(db: Session, inv: Invoice) -> None:
    total_alloc = (
        db.query(func.coalesce(func.sum(PaymentAllocation.amount), 0))
          .filter(PaymentAllocation.invoice_id == inv.id)
          .scalar() or 0
    )
    due = float(inv.amount_due or 0)
    ta  = float(total_alloc)
    if ta >= due:
        new_status = "paid"
    elif ta > 0:
        new_status = "partial"
    else:
        new_status = "chasing"

    update_vals = {"paid_amount": ta, "status": new_status}
    if new_status == "paid" and not inv.paid_at:
        update_vals["paid_at"] = datetime.utcnow()

    db.query(Invoice).filter(Invoice.id == inv.id).update(
        update_vals, synchronize_session=False
    )
def _status_expr(inv, remaining):
    """
    Return a CASE() that yields 'paid' | 'overdue' | 'open'
    using MySQL CURDATE() and DATE(inv.due_date) to avoid TZ drift.
    """
    today_sql = func.curdate()
    overdue = and_(inv.due_date != None, func.date(inv.due_date) < today_sql, remaining > 0)
    return case(
        (remaining <= 0, "paid"),
        (overdue, "overdue"),
        else_="open"
    )

@router.post("", response_model=InvoiceOut)
def create_invoice(
    p: InvoiceIn,
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    # customer must belong to this user
    cust = (
        db.query(Customer)
          .filter(Customer.id == p.customer_id, Customer.user_id == user.id)
          .first()
    )
    if not cust:
        raise HTTPException(404, "Unknown customer")

    inv_no = (p.invoice_number or "").strip()
    if not inv_no:
        raise HTTPException(400, "Invoice number is required")

    issue = datetime.fromisoformat(p.issue_date) if p.issue_date else datetime.utcnow()
    ttype = p.terms_type or cust.terms_type or "net_30"
    tdays = p.terms_days if (p.terms_type == "custom") else (cust.terms_days if cust.terms_type == "custom" else None)
    due = datetime.fromisoformat(p.due_date) if p.due_date else compute_due_date(issue, ttype, tdays)

    inv = Invoice(
        user_id=user.id,
        customer_id=p.customer_id,
        invoice_number=inv_no,
        amount_due=p.amount_due,
        currency=p.currency,
        issue_date=issue,
        terms_type=ttype,
        terms_days=tdays,
        due_date=due,
        status="chasing",
    )

    db.add(inv)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=f'Invoice number "{p.invoice_number}" is already in use for this customer.'
        )
    db.refresh(inv)

    return InvoiceOut(
        id=inv.id,
        customer_id=inv.customer_id,
        invoice_number=inv.invoice_number,
        amount_due=inv.amount_due,
        status=inv.status,
        due_date=(inv.due_date.isoformat() if inv.due_date else None),
        reconciliation_ref=inv.reconciliation_ref
    )

@router.get("")
def list_invoices(
    filter: Optional[str] = None,
    limit: int = 100,
    customer_id: Optional[int] = None,
    order: str = "desc",
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    q = db.query(Invoice).filter(Invoice.user_id == user.id)

    if customer_id:
        # ensure customer belongs to this user
        owned = db.query(Customer.id).filter(Customer.id == customer_id, Customer.user_id == user.id).first()
        if not owned:
            raise HTTPException(404, "Unknown customer")
        q = q.filter(Invoice.customer_id == customer_id)

    if filter == "overdue":
        q = q.filter(Invoice.status != "paid", Invoice.status != "written_off", Invoice.due_date < func.now())
    elif filter == "paid":
        q = q.filter(Invoice.status == "paid")
    # etc…

    q = q.order_by(Invoice.id.desc() if order.lower() == "desc" else Invoice.id.asc()).limit(limit)
    rows = q.all()

    return [
        {
            "id": r.id,
            "customer_id": r.customer_id,
            "invoice_number": r.invoice_number,
            "amount_due": float(r.amount_due or 0),
            "issue_date": r.issue_date.isoformat() if r.issue_date else None,
            "due_date": r.due_date.isoformat() if r.due_date else None,
            "status": r.status,
        }
        for r in rows
    ]

@router.get("/list", response_model=PageOut)
def list_invoices(
    status: str = "all",                 # all|open|overdue|paid
    page: int = 1,
    per_page: int = 50,                  # UI: 20/50/100
    search: Optional[str] = None,        # matches customer name or invoice number
    date_from: Optional[str] = None,     # YYYY-MM-DD (issue_date)
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    user = Depends(require_user),
):
    # clamp
    allowed = (20, 50, 100)
    if per_page not in allowed:
        per_page = min(allowed, key=lambda x: abs(x - per_page))
    page = max(1, page)

    # allocations per-invoice subquery
    alloc_sum_sq = (
        db.query(
            PaymentAllocation.invoice_id.label("inv_id"),
            func.coalesce(func.sum(PaymentAllocation.amount), 0.0).label("alloc_sum"),
        )
        .group_by(PaymentAllocation.invoice_id)
        .subquery()
    )

    inv = Invoice
    cust = Customer
    today_sql = func.curdate()

    remaining_expr = (inv.amount_due - func.coalesce(alloc_sum_sq.c.alloc_sum, 0.0))
    status_expr = _status_expr(inv, remaining_expr)
    days_overdue_expr = func.greatest(
        func.datediff(today_sql, func.date(inv.due_date)),
        0
    )

    # base query
    q = (
        db.query(
            inv.id,
            inv.customer_id,
            cust.name.label("customer_name"),
            inv.invoice_number,
            inv.issue_date,
            inv.due_date,
            inv.amount_due.label("amount"),
            remaining_expr.label("remaining"),
            status_expr.label("status_calc"),
            days_overdue_expr.label("days_overdue_calc"),
        )
        .join(cust, cust.id == inv.customer_id)
        .outerjoin(alloc_sum_sq, alloc_sum_sq.c.inv_id == inv.id)
        .filter(inv.user_id == user.id, cust.user_id == user.id)
    )

    # date filter (issue date)
    if date_from:
        q = q.filter(func.date(inv.issue_date) >= datetime.fromisoformat(date_from).date())
    if date_to:
        q = q.filter(func.date(inv.issue_date) <= datetime.fromisoformat(date_to).date())

    # search
    if search:
        like = f"%{search.strip()}%"
        q = q.filter(or_(cust.name.ilike(like), inv.invoice_number.ilike(like)))

    # status filter
    if status in ("open", "overdue", "paid"):
        q = q.filter(status_expr == status)

    # order newest first
    q = q.order_by(inv.issue_date.desc(), inv.id.desc())

    # total count for paging
    total = q.count()
    pages = max(1, ceil(total / per_page))
    if page > pages:
        page = pages

    rows = q.limit(per_page).offset((page - 1) * per_page).all()

    # build rows directly from SQL-calculated fields
    items: List[InvoiceRow] = []
    for r in rows:
        amt = float(r.amount or 0.0)
        rem = float(r.remaining or 0.0)
        items.append(InvoiceRow(
            id=r.id,
            customer_id=r.customer_id,
            customer_name=r.customer_name or f"Customer #{r.customer_id}",
            invoice_number=r.invoice_number,
            issue_date=r.issue_date.date().isoformat() if r.issue_date else None,
            due_date=r.due_date.date().isoformat() if r.due_date else None,
            amount=round(amt, 2),
            remaining=round(rem, 2),
            status=r.status_calc,
            days_overdue=int(r.days_overdue_calc or 0),
        ))
    return PageOut(items=items, page=page, per_page=per_page, total=total, pages=pages)

# ----------------------------
# BULK UPLOAD SUPPORT
# ----------------------------

@router.get("/bulk-fields")
def bulk_fields():
    return {
        "fields": [
            {"key": "customer_name", "label": "Customer name", "required": False, "hint": "Required in Per-row mode"},
            {"key": "invoice_number", "label": "Invoice number", "required": True},
            {"key": "amount_due", "label": "Amount due", "required": True, "hint": "e.g. 120.00 (currency symbol ok)"},
            {"key": "issue_date", "label": "Invoice date", "required": False, "hint": "ISO or specify date_format"},
            {"key": "due_date", "label": "Due date", "required": False, "hint": "If blank, we compute from terms"},
            {"key": "terms_type", "label": "Terms type", "required": False, "hint": "net_30 | net_60 | month_following | custom"},
            {"key": "terms_days", "label": "Terms days (if custom)", "required": False},
            {"key": "currency", "label": "Currency", "required": False, "hint": "Default GBP"},
            {"key": "email", "label": "Customer email (match/create)", "required": False},
            {"key": "phone", "label": "Customer phone (match/create)", "required": False},
        ],
        "mapping_help": "Map by column letter (A,B,C,…) or header text. Use default_customer_id for single-customer uploads.",
    }

class BulkUploadResult(BaseModel):
    ok: bool
    inserted: int
    skipped_duplicates: int
    skipped_missing_customer: int
    errors: List[str] = []
    dry_run: bool = False
    sample: Optional[List[Dict[str, Any]]] = None
    unknown_customers: Optional[List[str]] = None

# helpers reused below
def _letter_to_index(val: str) -> Optional[int]:
    if not isinstance(val, str): return None
    m = re.fullmatch(r"[A-Za-z]", val.strip())
    if not m: return None
    return ord(val.upper()) - ord("A")

def _clean_decimal(s: str) -> Decimal:
    if s is None:
        raise InvalidOperation("empty amount")
    raw = str(s).strip()
    raw = raw.replace(",", "").replace("£", "")
    return Decimal(raw)

def _parse_date(s: Optional[str], date_format: Optional[str]) -> Optional[datetime]:
    if not s: return None
    s = s.strip()
    if not s: return None
    try:
        return datetime.fromisoformat(s)
    except Exception:
        pass
    if date_format:
        try:
            return datetime.strptime(s, date_format)
        except Exception:
            pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    try:
        return datetime.fromisoformat(s.replace("/", "-"))
    except Exception:
        raise HTTPException(400, f"Unrecognised date '{s}'. Provide date_format like '%d/%m/%Y'.")

def _normalise_terms(tt: Optional[str]) -> Optional[str]:
    if not tt: return None
    t = tt.strip().lower().replace(" ", "").replace("-", "")
    if t in ("net30","n30","30","30days"): return "net_30"
    if t in ("net60","n60","60","60days"): return "net_60"
    if t in ("monthfollowing","eom+1","eonm","endofnextmonth"): return "month_following"
    if t in ("custom","other"): return "custom"
    if t in ("net_30","net_60","month_following"): return t
    return "custom"

def _resolve_index(mapping_value: Optional[str], headers: Optional[List[str]]) -> Optional[int]:
    if not mapping_value: return None
    idx = _letter_to_index(mapping_value)
    if idx is not None:
        return idx
    if headers:
        try:
            return headers.index(mapping_value)
        except ValueError:
            pass
    return None

def read_excel_rows(binary: bytes) -> List[List[str]]:
    wb = load_workbook(filename=BytesIO(binary), read_only=True, data_only=True)
    ws = wb.active
    out: List[List[str]] = []
    for r in ws.iter_rows(values_only=True):
        out.append([("" if v is None else str(v)).strip() for v in r])
    return out

def detect_delimiter(sample: str) -> str:
    if not sample:
        return ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "|", "\t"])
        return dialect.delimiter
    except Exception:
        pass
    lines = [ln for ln in sample.splitlines() if ln.strip()][:10]
    if not lines:
        return ","
    candidates = [",", ";", "|", "\t"]
    def count_outside_quotes(line: str, delim: str) -> int:
        in_quote = False; c = 0
        for ch in line:
            if ch == '"': in_quote = not in_quote
            elif ch == delim and not in_quote: c += 1
        return c
    def score(delim: str):
        per_line = [count_outside_quotes(ln, delim) for ln in lines]
        avg = sum(per_line) / len(per_line)
        var = statistics.pvariance(per_line) if len(per_line) > 1 else 0.0
        return (avg, -var)
    best = max(candidates, key=score)
    return best if score(best)[0] >= 1 else ","

def guess_has_header(first_row: List[str]) -> bool:
    def looks_number(x: str) -> bool:
        try:
            float(str(x).replace(",", "").replace("£", ""))
            return True
        except Exception:
            return False
    if not first_row:
        return True
    numeric = sum(1 for c in first_row if looks_number(c))
    return numeric <= (len(first_row) // 2)

def suggest_mapping(headers: List[str]) -> Dict[str, str]:
    if not headers:
        return {}
    low = [h.strip().lower() for h in headers]
    def find(*names):
        for n in names:
            if n in low:
                return headers[low.index(n)]
        return None
    return {
        "customer_name":  find("customer","customer name","client","client name","company"),
        "invoice_number": find("invoice","invoice no","invoice number","inv no","inv #","ref","reference"),
        "amount_due":     find("amount","amount due","total","value"),
        "issue_date":     find("date","invoice date","issue date","raised","raised date"),
        "due_date":       find("due","due date"),
        "currency":       find("currency","curr"),
        "terms_type":     find("terms","payment terms"),
        "terms_days":     find("terms days","days"),
        "email":          find("email","e-mail"),
        "phone":          find("phone","mobile","telephone"),
    }

def guess_date_format_from_rows(values: List[str]) -> Optional[str]:
    formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]
    score = {fmt: 0 for fmt in formats}
    for v in values[:50]:
        v = (v or "").strip()
        if not v: 
            continue
        for fmt in formats:
            try:
                datetime.strptime(v.replace(".", "/"), fmt)
                score[fmt] += 1
            except Exception:
                pass
    best = max(score, key=score.get) if score else None
    return best if score.get(best, 0) > 0 else None

@router.post("/bulk-inspect")
async def bulk_inspect(
    csv_file: UploadFile = File(..., description="CSV or Excel to inspect"),
    header: bool | None = Form(None),
    delimiter: str | None = Form(None),
    assign_mode: str = Form("per_row"),
    default_customer_id: int | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    raw = await csv_file.read()
    fname = (csv_file.filename or "").lower()

    if fname.endswith(".xlsx"):
        rows = read_excel_rows(raw)
        if not rows:
            return {
                "ok": True, "headers": [], "preview": [],
                "suggested_mapping": {}, "delimiter": ",",
                "has_header": True, "missing_customers": []
            }
        has_header = bool(header) if header is not None else guess_has_header(rows[0])
        headers = rows[0] if has_header else [f"Column {i+1}" for i in range(len(rows[0]))]
        data_rows = rows[1:] if has_header else rows
        used_delim = ","
    else:
        text = raw.decode("utf-8-sig", errors="replace")
        used_delim = delimiter or detect_delimiter(text[:2048])
        reader = csv.reader(io.StringIO(text), delimiter=used_delim)
        rows = list(reader)
        if not rows:
            return {
                "ok": True, "headers": [], "preview": [],
                "suggested_mapping": {}, "delimiter": used_delim or ",",
                "has_header": True, "missing_customers": []
            }
        has_header = bool(header) if header is not None else guess_has_header(rows[0])
        headers = rows[0] if has_header else [f"Column {i+1}" for i in range(len(rows[0]))]
        data_rows = rows[1:] if has_header else rows

    preview: list[dict[str, Any]] = []
    for r in data_rows[:10]:
        rowd: dict[str, Any] = {}
        for i, col in enumerate(headers):
            rowd[col] = r[i] if i < len(r) else ""
        preview.append(rowd)

    mapping = {k: v for k, v in suggest_mapping(headers).items() if v}

    date_format_guess: Optional[str] = None
    if has_header and mapping.get("issue_date"):
        col_idx = headers.index(mapping["issue_date"])
        samples = [r[col_idx] for r in data_rows[:50] if col_idx < len(r)]
        date_format_guess = guess_date_format_from_rows(samples)

    missing_customers: list[str] = []
    if assign_mode == "per_row" and has_header and mapping.get("customer_name"):
        ci = headers.index(mapping["customer_name"])
        seen: set[str] = set()
        for r in data_rows[:1000]:
            if ci < len(r):
                nm = (r[ci] or "").strip()
                if not nm or nm in seen:
                    continue
                seen.add(nm)
                exists = (
                    db.query(Customer.id)
                      .filter(Customer.user_id == user.id)
                      .filter(func.lower(Customer.name) == nm.lower())
                      .first()
                )
                if not exists:
                    missing_customers.append(nm)
            if len(missing_customers) >= 50:
                break

    return {
        "ok": True,
        "delimiter": used_delim or ",",
        "has_header": has_header,
        "headers": headers,
        "preview": preview,
        "suggested_mapping": mapping,
        "date_format_guess": date_format_guess,
        "assign_mode": assign_mode,
        "missing_customers": missing_customers,
    }

@router.post("/bulk-upload", response_model=BulkUploadResult)
async def bulk_upload_multipart(
    csv_file: UploadFile = File(..., description="CSV or Excel file"),
    mapping: str = Form(..., description="JSON map of logical->column"),
    header: bool = Form(True),
    delimiter: str = Form(","),
    date_format: Optional[str] = Form(None),
    create_missing_customers: bool = Form(False),
    dry_run: bool = Form(False),
    default_customer_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    try:
        mapping_obj = json.loads(mapping or "{}")
        if not isinstance(mapping_obj, dict):
            raise ValueError
    except Exception:
        raise HTTPException(400, "Invalid mapping JSON")

    raw = await csv_file.read()
    fname = (csv_file.filename or "").lower()

    if fname.endswith(".xlsx"):
        rows = read_excel_rows(raw)
        if not rows:
            return BulkUploadResult(ok=True, inserted=0, skipped_duplicates=0,
                                    skipped_missing_customer=0, errors=["Empty file"])
        headers = rows[0] if header else None
        data_rows = rows[1:] if header else rows
    else:
        text = raw.decode("utf-8-sig", errors="replace")
        use_delim = delimiter or detect_delimiter(text[:65536])
        reader = csv.reader(io.StringIO(text), delimiter=use_delim)
        rows = list(reader)
        if not rows:
            return BulkUploadResult(ok=True, inserted=0, skipped_duplicates=0,
                                    skipped_missing_customer=0, errors=["Empty file"])
        headers = rows[0] if header else None
        data_rows = rows[1:] if header else rows

    def _resolve_index(mapping_value: Optional[str], headers: Optional[List[str]]) -> Optional[int]:
        if not mapping_value: return None
        # letter?
        if isinstance(mapping_value, str) and len(mapping_value.strip()) == 1 and mapping_value.isalpha():
            return ord(mapping_value.upper()) - ord("A")
        # header text?
        if headers:
            try:
                return headers.index(mapping_value)
            except ValueError:
                pass
        return None

    idx = {k: _resolve_index(v, headers) for k, v in mapping_obj.items()}

    for r in ["invoice_number", "amount_due"]:
        if idx.get(r) is None:
            raise HTTPException(400, f"Mapping missing or invalid for required field '{r}'.")

    per_row_mode = (default_customer_id is None)
    if per_row_mode and idx.get("customer_name") is None:
        raise HTTPException(400, "Per-row mode: map the 'Customer name' column or select a single customer.")

    inserted = 0
    skipped_dup = 0
    skipped_missing_cust = 0
    errors: List[str] = []
    sample_out: List[Dict[str, Any]] = []
    unknown: Set[str] = set()

    cust_cache: Dict[str, Optional[int]] = {}

    def find_or_create_customer(name: Optional[str], email: Optional[str], phone: Optional[str],
                                default_id: Optional[int]) -> Optional[int]:
        if default_id:
            # verify default belongs to user
            owned = db.query(Customer.id).filter(Customer.id == default_id, Customer.user_id == user.id).first()
            return default_id if owned else None

        if email:
            c = (db.query(Customer)
                  .filter(Customer.user_id == user.id)
                  .filter(func.lower(Customer.email) == email.strip().lower())
                  .first())
            if c: return c.id

        if phone:
            norm = re.sub(r"\D+", "", phone)
            if norm:
                for c in db.query(Customer).filter(Customer.user_id == user.id).all():
                    if re.sub(r"\D+", "", c.phone or "") == norm:
                        return c.id

        if name:
            key = (name or "").strip().lower()
            if key in cust_cache:
                return cust_cache[key]
            c = (db.query(Customer)
                  .filter(Customer.user_id == user.id)
                  .filter(func.lower(Customer.name) == key)
                  .first())
            if c:
                cust_cache[key] = c.id
                return c.id

        if create_missing_customers and name:
            c = Customer(user_id=user.id, name=name.strip(), email=(email or None), phone=(phone or None))
            db.add(c); db.flush()
            if name:
                cust_cache[(name or "").strip().lower()] = c.id
            return c.id
        return None

    for r_idx, row in enumerate(data_rows, start=(2 if header else 1)):
        try:
            def get(k: str) -> Optional[str]:
                j = idx.get(k)
                if j is None: return None
                if j < 0 or j >= len(row): return None
                val = (row[j] or "").strip()
                return val if val != "" else None

            cust_name = get("customer_name")
            inv_no    = get("invoice_number")
            amt_raw   = get("amount_due")

            if not inv_no or amt_raw is None or (per_row_mode and not cust_name):
                errors.append(f"Row {r_idx}: missing required fields")
                continue

            currency  = get("currency") or "GBP"
            issue_s   = get("issue_date")
            due_s     = get("due_date")
            terms_t   = _normalise_terms(get("terms_type"))
            terms_d_s = get("terms_days")
            email     = get("email")
            phone     = get("phone")

            customer_id = find_or_create_customer(cust_name, email, phone, default_customer_id)
            if not customer_id:
                skipped_missing_cust += 1
                if cust_name:
                    unknown.add(cust_name.strip())
                continue

            exists = (
                db.query(Invoice.id)
                  .filter(Invoice.user_id == user.id, Invoice.customer_id == customer_id, Invoice.invoice_number == inv_no)
                  .first()
            )
            if exists:
                skipped_dup += 1
                continue

            amount = _clean_decimal(amt_raw)

            issue_dt = _parse_date(issue_s, date_format) or datetime.utcnow()
            if due_s:
                due_dt = _parse_date(due_s, date_format)
            else:
                cust_obj = db.query(Customer).filter(Customer.id == customer_id, Customer.user_id == user.id).first()
                ttype = terms_t or (cust_obj.terms_type if cust_obj and cust_obj.terms_type else "net_30")
                tdays = None
                if (terms_t == "custom" and terms_d_s):
                    try:
                        tdays = int(terms_d_s)
                    except Exception:
                        tdays = None
                elif cust_obj and cust_obj.terms_type == "custom":
                    tdays = cust_obj.terms_days
                due_dt = compute_due_date(issue_dt, ttype, tdays)

            out_row = {
                "customer_id": customer_id,
                "customer_name": cust_name,
                "invoice_number": inv_no,
                "amount_due": str(amount),
                "currency": currency,
                "issue_date": issue_dt.isoformat(),
                "due_date": due_dt.isoformat() if due_dt else None,
            }

            if dry_run:
                if len(sample_out) < 10:
                    sample_out.append(out_row)
                continue

            inv = Invoice(
                user_id=user.id,
                customer_id=customer_id,
                invoice_number=inv_no,
                amount_due=amount,
                currency=currency,
                issue_date=issue_dt,
                due_date=due_dt,
                status="chasing",
                terms_type=terms_t or None,
                terms_days=int(terms_d_s) if (terms_t == "custom" and terms_d_s and str(terms_d_s).isdigit()) else None,
            )
            db.add(inv)
            inserted += 1

        except InvalidOperation:
            errors.append(f"Row {r_idx}: invalid amount '{amt_raw}'")
        except HTTPException as he:
            errors.append(f"Row {r_idx}: {he.detail}")
        except Exception as ex:
            errors.append(f"Row {r_idx}: {type(ex).__name__}: {ex}")

    if dry_run:
        db.rollback()
        return BulkUploadResult(
            ok=True,
            inserted=0,
            skipped_duplicates=0,
            skipped_missing_customer=skipped_missing_cust,
            errors=errors,
            dry_run=True,
            sample=sample_out,
            unknown_customers=sorted(unknown) if unknown else [],
        )

    db.commit()
    return BulkUploadResult(
        ok=True,
        inserted=inserted,
        skipped_duplicates=skipped_dup,
        skipped_missing_customer=skipped_missing_cust,
        errors=errors,
        dry_run=False,
        unknown_customers=sorted(unknown) if unknown else None,
    )

# ----------------------------
# Upload Presets (save/load mappings & options)
# ----------------------------

@router.get("/upload-presets")
def list_presets(db: Session = Depends(get_db)):
    presets = db.query(InvoiceUploadPreset).order_by(InvoiceUploadPreset.created_at.desc()).all()
    return [{"id": p.id, "name": p.name} for p in presets]

@router.get("/upload-presets/{preset_id}", response_model=UploadPresetOut)
def get_preset(preset_id: int, db: Session = Depends(get_db)):
    p = db.get(InvoiceUploadPreset, preset_id)
    if not p:
        raise HTTPException(404, "Preset not found")
    return UploadPresetOut(
        id=p.id, name=p.name,
        mapping=json.loads(p.mapping_json or "{}"),
        header=bool(p.header),
        delimiter=p.delimiter or ",",
        date_format=p.date_format,
        default_customer_id=p.default_customer_id,
        assign_mode=p.assign_mode or "per_row",
        create_missing_customers=bool(p.create_missing_customers),
    )

@router.post("/upload-presets")
def create_preset(p: UploadPresetIn, db: Session = Depends(get_db)):
    exists = db.query(InvoiceUploadPreset).filter(InvoiceUploadPreset.name == p.name).first()
    if exists:
        raise HTTPException(409, "A preset with this name already exists")

    preset = InvoiceUploadPreset(
        name=p.name,
        mapping_json=json.dumps(p.mapping or {}),
        header=p.header,
        delimiter=p.delimiter or ",",
        date_format=p.date_format,
        default_customer_id=p.default_customer_id,
        assign_mode=p.assign_mode if p.assign_mode in ("per_row","single") else "per_row",
        create_missing_customers=p.create_missing_customers,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(preset); db.commit(); db.refresh(preset)
    return {"id": preset.id, "name": preset.name}

@router.delete("/upload-presets/{preset_id}")
def delete_preset(preset_id: int, db: Session = Depends(get_db)):
    p = db.get(InvoiceUploadPreset, preset_id)
    if not p:
        raise HTTPException(404, "Preset not found")
    db.delete(p); db.commit()
    return {"ok": True}
